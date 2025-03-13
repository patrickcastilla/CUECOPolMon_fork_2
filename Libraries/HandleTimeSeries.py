from pandas import DataFrame, DatetimeIndex, Timestamp, Period, to_datetime, date_range
from numpy import isnan
from math import floor, ceil
import matplotlib.pyplot as plt
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from pandas.tseries.offsets import MonthEnd, QuarterEnd
import re
from sympy import Symbol, Matrix

python_version = sys.version_info
#python_version.major
#python_version.minor

def quarterly_dates_from_end(end_quarter_str, num_periods):
    try:
        end_period = Period(end_quarter_str)
    except ValueError:
        raise ValueError("Invalid end_quarter_str.  Must be in YYYYQ format (e.g., '2023Q4').")

    if not isinstance(num_periods, int) or num_periods <= 0:
        raise ValueError("num_periods must be a positive integer.")

    start_period = end_period - (num_periods-1)
    if python_version.minor >=12:
        dates = date_range(start=start_period.to_timestamp(), periods=num_periods, freq='QE')
    else:
        dates = date_range(start=start_period.to_timestamp(), periods=num_periods, freq='Q')

    return dates

def generate_quarterly_dates(start_date, horizon_quarters):
    if not isinstance(start_date, (str, Timestamp, Period, DatetimeIndex)):
        try:
            start_date = to_datetime(start_date) # Try converting if it is something that can be converted
        except:
            raise TypeError("start_date must be a string or datetime-like object")
            
    if not isinstance(horizon_quarters, int):
        raise TypeError("horizon_quarters must be an integer")

    if horizon_quarters < 0:
        raise ValueError("horizon_quarters must be non-negative")

    try:
        start_date = to_datetime(start_date) # Convert to datetime if it is not already
        if python_version.minor >=12:
            quarterly_dates = date_range(start=start_date, periods=horizon_quarters + 1, freq='QE')
        else:
            quarterly_dates = date_range(start=start_date, periods=horizon_quarters + 1, freq='Q')
        return quarterly_dates

    except Exception as e:  # Catch potential errors during date creation
        print(f"An error occurred: {e}")
        return None
    
def get_EndHistory(DFHistory):
    firstendnotfound = True
    index = DFHistory.index
    pos = -1
    while firstendnotfound:
        if isnan(DFHistory.iloc[pos]).any():
            pos += -1
        else:
            firstendnotfound = False
        
    return index[pos]

def get_FirstHistory(DFHistory):
    lastbeginningnotfound = True
    index = DFHistory.index
    pos = 1
    while lastbeginningnotfound:
        if isnan(DFHistory.iloc[pos]).any():
            pos += 1
        else:
            lastbeginningnotfound = False
        
    return index[pos-1]


def plotTSforecast(DBdict,lista,listaname,SS,FH,HH=None,nr=None,nc=None,SS_plot = True):
    
    nv = len(lista)
    if HH is None:
        HH = 12

    if (nr is None) and (nc is None):
        nr = floor(nv**(0.5))
        nc = nr+1
    
    if (nr is not None) and (nc is None):
        nc = ceil(nv/nr)

    if (nr is None) and (nc is not None):
        nc = ceil(nv/nr)

    fig, axs = plt.subplots(nr,nc, figsize=(nc*4, nr * 3))

    for ii in range(nv):
        ck = -1
        for key in DBdict:
            ck += 1
            DB = DBdict[key]
            if (nr==1) or (nc==1):
                if ck==0:
                    axs[ii].plot(DB.index[-(FH+HH):-FH], DB[lista[ii]][-(FH+HH):-FH], linestyle='-',marker='o',label="Hist.")
                    axs[ii].plot(DB.index[-(FH+1):], DB[lista[ii]][-(FH+1):], linestyle='--',label=str(key))
                else:
                    axs[ii].plot(DB.index[-(FH+1):], DB[lista[ii]][-(FH+1):], linestyle='--',label=str(key))

                if SS_plot:
                    axs[ii].axhline(y=SS[ii], color='black', linestyle='-.', linewidth=0.75)
                axs[ii].axvline(x=DB.index[-(FH+1)], color='black', linestyle='-.', linewidth=0.75)
                axs[ii].set_title(listaname[ii])
            else:
                nri = floor(ii/nc)
                nci = ii - floor(ii/nc)*nc
                if ck==0:
                    axs[nri,nci].plot(DB.index[-(FH+HH):-FH], DB[lista[ii]][-(FH+HH):-FH], linestyle='-',marker='o',label="Hist.")
                    axs[nri,nci].plot(DB.index[-(FH+1):], DB[lista[ii]][-(FH+1):], linestyle='--',label=str(key))
                else:
                    axs[nri,nci].plot(DB.index[-(FH+1):], DB[lista[ii]][-(FH+1):], linestyle='--',label=str(key))
                if SS_plot:
                    axs[nri,nci].axhline(y=SS[ii], color='black', linestyle='-.', linewidth=0.75)
                axs[nri,nci].axvline(x=DB.index[-(FH+1)], color='black', linestyle='-.', linewidth=0.75)
                axs[nri,nci].set_title(listaname[ii])
        if ii==0:
            axs[nri,nci].legend(loc="best")
        
    plt.tight_layout()

    return fig, axs


def DB2excel(filename, dataframe, sheet="Data", Description=None):
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        dataframe.to_excel(
            writer,
            sheet_name=sheet,
            startrow=2,
            startcol=0,
            index=True,
            header=False,
        )

        worksheet = writer.sheets[sheet]
        workbook = writer.book

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])


        for idx, col_name in enumerate(dataframe.columns):
            worksheet.cell(row=1, column=idx + 2, value=col_name)

        if Description is not None:
            for idx, desc in enumerate(Description):
                worksheet.cell(row=2, column=idx + 2, value=desc)


        no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
        font_normal = Font(bold=False)
        
        for row in range(3, len(dataframe) + 3):
            cell = worksheet.cell(row=row, column=1)
            cell.font = font_normal
            cell.border = no_border
            cell.number_format = 'dd/mm/yyyy' 

        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    value = str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
            
            
def ReadDB_excel(filename, sheet='Data'):
    header_df = pd.read_excel(filename, sheet_name=sheet, header=None, nrows=2)
    
    diccionario_de_descripciones = {}
    for col in range(1, header_df.shape[1]):
        nombre_col = header_df.iloc[0, col]
        descripcion = header_df.iloc[1, col]
        diccionario_de_descripciones[nombre_col] = descripcion
    df = pd.read_excel(filename, sheet_name=sheet, header=0, skiprows=[1], index_col=0)
    
    return diccionario_de_descripciones, df



def ReadForeCond(filename, f0, FH):
    start_date = pd.to_datetime(f0)
    quarterly_dates = generate_quarterly_dates(f0, FH-1)
    # Después (correcto)
    if quarterly_dates.empty:  # <-- Usar .empty para verificar
       raise ValueError("Could not generate the quarterly date index")
    end_date = quarterly_dates[-1]
    adjusted_start_date = start_date + QuarterEnd(startingMonth=3)
    
    xls = pd.ExcelFile(filename)
    objeto = {}
    
    for sheet in xls.sheet_names:
        # Eliminar ceros finales en el nombre de la hoja
        cleaned_sheet_name = sheet.rstrip('0') if sheet.endswith('0') else sheet
        
        df_sheet = pd.read_excel(xls, sheet_name=sheet, header=None)
        
        # Caso sin shock
        if df_sheet.iloc[1, 1:].isna().all():
            headers = df_sheet.iloc[0, 1:].tolist()
            data_rows = df_sheet.iloc[1:].dropna(subset=[0])
            
            if not data_rows.empty:
                # Procesar fechas y frecuencia
                orig_index = pd.to_datetime(data_rows[0]) + MonthEnd(0)
                detected_freq = pd.infer_freq(orig_index) or _infer_freq_manual(orig_index)
                new_index = _generate_new_index(adjusted_start_date, end_date, detected_freq)
                
                # Crear DataFrame
                data = data_rows.iloc[:, 1:].set_axis(headers, axis=1)
                df_final = data.set_index(orig_index).reindex(new_index)
            else:
                df_final = pd.DataFrame(columns=headers)
            
            objeto[cleaned_sheet_name] = df_final
        
        # Caso con shock
        else:
            headers = df_sheet.iloc[0, 1:].tolist()
            shock = df_sheet.iloc[1, 1:].tolist()
            data_rows = df_sheet.iloc[2:].dropna(subset=[0])
            
            if data_rows.empty:
                df_final = pd.DataFrame(columns=headers)
            else:
                orig_index = pd.to_datetime(data_rows[0]) + MonthEnd(0)
                detected_freq = pd.infer_freq(orig_index) or _infer_freq_manual(orig_index)
                new_index = _generate_new_index(adjusted_start_date, end_date, detected_freq)
                
                data = data_rows.iloc[:, 1:].set_axis(headers, axis=1)
                df_final = data.set_index(orig_index).reindex(new_index)
            
            objeto[cleaned_sheet_name] = {"var": df_final, "shock": shock}
    objeto = convertir_a_float(objeto)
    return objeto

# Funciones auxiliares
def _infer_freq_manual(index):
    if len(index) >= 2:
        diff = index[1] - index[0]
        return pd.tseries.frequencies.to_offset(diff).freqstr
    return None

def _generate_new_index(start, end, freq):
    new_index = pd.date_range(start=start, end=end, freq=freq)
    if new_index[0] != start:
        new_index = new_index.insert(0, start)
    return new_index + MonthEnd(0)


def convertir_a_float(estructura):
    """Convierte todos los valores numéricos en una estructura anidada a float."""
    if isinstance(estructura, dict):
        # Procesar diccionarios recursivamente
        return {k: convertir_a_float(v) for k, v in estructura.items()}
    elif isinstance(estructura, pd.DataFrame):
        # Convertir DataFrame completo a float
        return estructura.astype(float)
    elif isinstance(estructura, list):
        # Procesar listas recursivamente
        return [convertir_a_float(item) for item in estructura]
    else:
        # Dejar otros tipos sin cambios
        return estructura


def transformar_nombre_variable(var_name: str) -> str:
    count = var_name.count('Ͱ')  
    if count == 1:

        match_1_1 = re.match(r'^l(\d+)Ͱ([\w\d_]+)$', var_name)
        if match_1_1:
            numero = int(match_1_1.group(1))
            var = match_1_1.group(2)
            return f"{var}(-{numero})"

        match_1_2 = re.match(r'^([\w\d_]+)Ͱ(\d+)p$', var_name)
        if match_1_2:
            var = match_1_2.group(1)
            numero = int(match_1_2.group(2))
            return f"{var}(+{numero})"
        return var_name

    elif count == 2:
        match_2 = re.match(r'^l(\d+)Ͱ([\w\d_]+)Ͱ(\d+)p$', var_name)
        if match_2:
            numero1 = int(match_2.group(1))
            var = match_2.group(2)
            numero2 = int(match_2.group(3))
            
            if numero1 > numero2:
                return f"{var}(-{numero1 - numero2})"
            elif numero1 < numero2:
                return f"{var}(+{numero2 - numero1})"
            else:
                return f"{var}"
        
        return var_name

    else:
        return var_name


def transformar_matriz_varnames(var_matrix):
    var_matrix = Matrix(var_matrix)
    filas, cols = var_matrix.shape
    nueva_matriz = Matrix.zeros(filas, cols)

    for i in range(filas):
        for j in range(cols):
            original_name = str(var_matrix[i, j])
            nuevo_nombre = transformar_nombre_variable(original_name)
            nuevo_simbolo = Symbol(nuevo_nombre, real=True)
            nueva_matriz[i, j] = nuevo_simbolo

    return nueva_matriz